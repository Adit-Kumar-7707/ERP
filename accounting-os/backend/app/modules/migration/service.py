from sqlalchemy.orm import Session
from fastapi import UploadFile, HTTPException
import openpyxl
import csv
import io
from typing import List, Dict, Any
from datetime import date
from json import dumps

from app.modules.migration.models import ImportBatch, ImportLog, ImportStatus, ImportType
from app.modules.accounting.models import Account, AccountType, JournalEntry, JournalLine
from app.modules.organization.models import Organization

def validate_row_schema(row: Dict, required: List[str]) -> List[str]:
    errors = []
    for field in required:
        if field not in row or not row[field]:
             # Allow 0 for numeric
            if isinstance(row.get(field), (int, float)) and row.get(field) == 0:
                continue
            errors.append(f"Missing {field}")
    return errors

def process_excel_import(db: Session, file: UploadFile, user_id: int):
    # 1. Create Batch
    batch = ImportBatch(
        filename=file.filename, 
        type=ImportType.EXCEL, 
        imported_by_id=user_id,
        status=ImportStatus.PENDING
    )
    db.add(batch)
    db.commit()
    db.refresh(batch)
    
    try:
        # 2. Parse Data
        contents = file.file.read()
        rows = []
        
        if file.filename.endswith(".csv"):
             # CSV parsing
             decoded = contents.decode("utf-8-sig")
             reader = csv.DictReader(io.StringIO(decoded))
             rows = list(reader)
        else:
             # Excel parsing
             wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
             ws = wb.active
             headers = [cell.value for cell in ws[1]]
             for row in ws.iter_rows(min_row=2, values_only=True):
                 rows.append(dict(zip(headers, row)))

        batch.record_count = len(rows)
        
        # 3. Validate & Accumulate
        # Expected Schema: Name, Group, Debit, Credit
        ledger_entries = []
        total_dr = 0.0
        total_cr = 0.0
        
        logs = []
        has_error = False
        
        for idx, row in enumerate(rows):
            # Clean keys (strip spaces)
            row = {k.strip(): v for k, v in row.items() if k}
            
            # Validation
            validation_errs = validate_row_schema(row, ["Name", "Group"])
            # Dr/Cr optional but one must be float
            dr = float(row.get("Debit") or 0)
            cr = float(row.get("Credit") or 0)
            
            if dr < 0 or cr < 0:
                validation_errs.append("Negative values not allowed")
                
            if validation_errs:
                has_error = True
                logs.append(ImportLog(batch_id=batch.id, row_number=idx+2, status="Failed", error_message="; ".join(validation_errs), raw_data=dumps(str(row))))
                continue
            
            ledger_entries.append({
                "name": row["Name"],
                "group": row["Group"], # Tally Group or System Group? Assuming Name matches System Group or we create it?
                                      # Simplification: Assume System Group Name is provided.
                "debit": dr,
                "credit": cr
            })
            total_dr += dr
            total_cr += cr
            
        # 4. Check Balance
        if abs(total_dr - total_cr) > 0.01:
             has_error = True
             logs.append(ImportLog(batch_id=batch.id, row_number=0, status="Failed", error_message=f"Trial Balance Mismatch: Dr {total_dr} != Cr {total_cr}", raw_data=None))
        
        if has_error:
             batch.status = ImportStatus.FAILED
             db.add_all(logs)
             db.commit()
             return {"status": "Failed", "errors": [l.error_message for l in logs]}
             
        # 5. Transactional Import
        # A. Create Journal Entry Header
        # B. Create Accounts (if missing)
        # C. Create Lines
        
        # We need a Financial Year? Opening usually first day of active FY.
        # Fetch Active FY
        from app.modules.accounting.models import FinancialYear
        fy = db.query(FinancialYear).filter(FinancialYear.is_active == True).first()
        if not fy:
             raise Exception("No Active Financial Year found")
             
        date = fy.start_date
        
        journal_entry = JournalEntry(
            date=date,
            voucher_type="Journal",
            reference="OB-IMPORT",
            narration=f"Opening Balance Import from {file.filename}",
            financial_year_id=fy.id,
            is_opening=True
        )
        db.add(journal_entry)
        db.flush()
        
        # Cache Groups
        groups = db.query(Account).filter(Account.is_group == True).all()
        group_map = {g.name.lower(): g.id for g in groups}
        
        # Fallback root groups
        roots = {
            "assets": "AST", "liabilities": "LIA", "income": "INC", "expenses": "EXP", "equity": "EQ"
        }
        
        for entry in ledger_entries:
            # Check if Account exists
            acc = db.query(Account).filter(Account.name == entry["name"]).first()
            if not acc:
                # Create Account
                # Find Group ID
                grp_name = entry["group"]
                pid = group_map.get(grp_name.lower())
                
                # If group missing, fail? Or create? 
                # Checklist: "Import with missing group: MUST FAIL".
                if not pid:
                     raise Exception(f"Group '{grp_name}' not found for Ledger '{entry['name']}'")
                
                # Create
                code = f"{entry['name'][:3].upper()}-{date}" # Simple code gen
                
                # Determine Type from Group? 
                # Account needs Type. Group has Type? 
                # We need to fetch Group object to get Type? 
                # Impl detail: Account model: `type`. Is derived from Parent?
                # Yes, typically.
                parent = db.query(Account).filter(Account.id == pid).first()
                
                acc = Account(
                    name=entry["name"],
                    code=code + str(idx), # Unique
                    type=parent.type,
                    parent_id=pid
                )
                db.add(acc)
                db.flush()
            
            # Create Line
            if entry["debit"] > 0:
                 db.add(JournalLine(entry_id=journal_entry.id, account_id=acc.id, debit=entry["debit"], credit=0))
            if entry["credit"] > 0:
                 db.add(JournalLine(entry_id=journal_entry.id, account_id=acc.id, debit=0, credit=entry["credit"]))
                 
        batch.status = ImportStatus.SUCCESS
        db.commit()
        return {"status": "Success", "records": len(ledger_entries)}
        
    except Exception as e:
        db.rollback()
        batch.status = ImportStatus.FAILED
        # Log generic error
        db.add(ImportLog(batch_id=batch.id, status="Failed", error_message=str(e)))
        db.commit()
        raise HTTPException(status_code=400, detail=f"Import Failed: {str(e)}")
